"""
test_consolidation.py — Testes do plano de consolidação 301 (P2).

Cobertura (lógica pura, sem IO de rede):
  - build_consolidation_plan: escolha da canônica (cliques desc → posição asc
    → impressões desc), conflitos entre grupos, ausência de cadeias/ciclos
  - build_htaccess_block / build_nginx_block: formato e aviso de SUGESTÃO
  - storage.save_redirects_csv / save_redirects_txt: colunas e conteúdo
"""

import csv
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core import storage
from core.analytics import (
    build_consolidation_plan,
    build_htaccess_block,
    build_nginx_block,
)


def _url(d: dict) -> dict:
    """URL de grupo de canibalização com defaults."""
    base = {
        "url": "https://ex.com/a",
        "position": 5.0,
        "clicks": 10,
        "impressions": 100,
        "ctr": 5.0,
    }
    base.update(d)
    return base


def _group(query="kw", severity="alta", urls=None) -> dict:
    return {
        "query": query,
        "severity": severity,
        "url_count": len(urls or []),
        "urls": urls or [],
        "severity_score": 0,
    }


# ---------------------------------------------------------------------------
# Escolha da canônica
# ---------------------------------------------------------------------------


class TestCanonicalPick(unittest.TestCase):
    def test_mais_cliques_vence(self):
        g = _group(
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 5, "position": 1.0}),
                _url({"url": "https://ex.com/b", "clicks": 50, "position": 9.0}),
            ]
        )
        plan = build_consolidation_plan([g])
        self.assertEqual(plan["groups"][0]["canonical"]["url"], "https://ex.com/b")
        self.assertEqual(plan["redirects"][0]["from_url"], "https://ex.com/a")
        self.assertEqual(plan["redirects"][0]["to_url"], "https://ex.com/b")

    def test_empate_cliques_melhor_posicao_vence(self):
        g = _group(
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 10, "position": 8.0}),
                _url({"url": "https://ex.com/b", "clicks": 10, "position": 3.0}),
            ]
        )
        plan = build_consolidation_plan([g])
        self.assertEqual(plan["groups"][0]["canonical"]["url"], "https://ex.com/b")

    def test_empate_total_mais_impressoes_vence(self):
        g = _group(
            urls=[
                _url(
                    {"url": "https://ex.com/a", "clicks": 10, "position": 5.0, "impressions": 100}
                ),
                _url(
                    {"url": "https://ex.com/b", "clicks": 10, "position": 5.0, "impressions": 900}
                ),
            ]
        )
        plan = build_consolidation_plan([g])
        self.assertEqual(plan["groups"][0]["canonical"]["url"], "https://ex.com/b")

    def test_grupo_com_3_urls_gera_2_redirects(self):
        g = _group(
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 50}),
                _url({"url": "https://ex.com/b", "clicks": 5}),
                _url({"url": "https://ex.com/c", "clicks": 1}),
            ]
        )
        plan = build_consolidation_plan([g])
        self.assertEqual(plan["total_redirects"], 2)
        self.assertEqual(
            {r["from_url"] for r in plan["redirects"]}, {"https://ex.com/b", "https://ex.com/c"}
        )
        self.assertTrue(all(r["to_url"] == "https://ex.com/a" for r in plan["redirects"]))

    def test_campos_do_redirect(self):
        g = _group(
            query="cane corso preço",
            severity="alta",
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 50}),
                _url({"url": "https://ex.com/b", "clicks": 7}),
            ],
        )
        r = build_consolidation_plan([g])["redirects"][0]
        self.assertEqual(r["keyword"], "cane corso preço")
        self.assertEqual(r["severity"], "alta")
        self.assertEqual(r["clicks_from"], 7)
        self.assertEqual(r["clicks_to"], 50)

    def test_plano_vazio(self):
        plan = build_consolidation_plan([])
        self.assertEqual(plan["total_redirects"], 0)
        self.assertEqual(plan["groups"], [])
        self.assertIn("SUGESTAO", plan["disclaimer"])


# ---------------------------------------------------------------------------
# Conflitos entre grupos
# ---------------------------------------------------------------------------


class TestConflicts(unittest.TestCase):
    def test_canonica_de_um_grupo_nao_vira_fonte_em_outro(self):
        # B é canônica no grupo 1 (mais severo); grupo 2 quer redirecionar B
        g1 = _group(
            query="kw1",
            urls=[
                _url({"url": "https://ex.com/b", "clicks": 50}),
                _url({"url": "https://ex.com/a", "clicks": 5}),
            ],
        )
        g2 = _group(
            query="kw2",
            urls=[
                _url({"url": "https://ex.com/c", "clicks": 90}),
                _url({"url": "https://ex.com/b", "clicks": 50}),
            ],
        )
        plan = build_consolidation_plan([g1, g2])
        from_urls = {r["from_url"] for r in plan["redirects"]}
        self.assertNotIn("https://ex.com/b", from_urls)
        self.assertTrue(any("kw2" in c for c in plan["conflicts"]))

    def test_fonte_nao_recebe_segundo_destino(self):
        # A redireciona para B no grupo 1; grupo 2 tentaria A → C
        g1 = _group(
            query="kw1",
            urls=[
                _url({"url": "https://ex.com/b", "clicks": 50}),
                _url({"url": "https://ex.com/a", "clicks": 5}),
            ],
        )
        g2 = _group(
            query="kw2",
            urls=[
                _url({"url": "https://ex.com/c", "clicks": 90}),
                _url({"url": "https://ex.com/a", "clicks": 5}),
            ],
        )
        plan = build_consolidation_plan([g1, g2])
        a_redirects = [r for r in plan["redirects"] if r["from_url"] == "https://ex.com/a"]
        self.assertEqual(len(a_redirects), 1)
        self.assertEqual(a_redirects[0]["to_url"], "https://ex.com/b")
        self.assertTrue(any("destino mantido" in c for c in plan["conflicts"]))

    def test_url_ja_redirecionada_nao_candidata_a_canonica(self):
        # A → B no grupo 1. No grupo 2, A teria mais cliques, mas já foi
        # redirecionada — a canônica do grupo 2 deve ser outra URL.
        g1 = _group(
            query="kw1",
            urls=[
                _url({"url": "https://ex.com/b", "clicks": 100}),
                _url({"url": "https://ex.com/a", "clicks": 90}),
            ],
        )
        g2 = _group(
            query="kw2",
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 90}),
                _url({"url": "https://ex.com/c", "clicks": 10}),
            ],
        )
        plan = build_consolidation_plan([g1, g2])
        g2_out = [g for g in plan["groups"] if g["query"] == "kw2"]
        if g2_out:  # grupo só entra no plano se gerou redirect
            self.assertEqual(g2_out[0]["canonical"]["url"], "https://ex.com/c")

    def test_sem_cadeias_nem_ciclos(self):
        # Propriedade estrutural: nenhum from_url aparece como to_url
        groups = [
            _group(
                query=f"kw{i}",
                urls=[
                    _url({"url": f"https://ex.com/{i}a", "clicks": 50 - i}),
                    _url({"url": f"https://ex.com/{i}b", "clicks": 5}),
                    _url({"url": f"https://ex.com/{(i + 1) % 3}a", "clicks": 20}),
                ],
            )
            for i in range(3)
        ]
        plan = build_consolidation_plan(groups)
        from_urls = {r["from_url"] for r in plan["redirects"]}
        to_urls = {r["to_url"] for r in plan["redirects"]}
        self.assertEqual(from_urls & to_urls, set(), "plano nao pode conter cadeias de redirect")


# ---------------------------------------------------------------------------
# Blocos de servidor (Apache / nginx)
# ---------------------------------------------------------------------------


def _plan_simples() -> dict:
    g = _group(
        query="kw açaí",
        severity="média",
        urls=[
            _url({"url": "https://www.ex.com.br/melhor-pagina/", "clicks": 50}),
            _url({"url": "https://www.ex.com.br/pagina-dup/?p=1", "clicks": 5}),
        ],
    )
    return build_consolidation_plan([g])


class TestServerBlocks(unittest.TestCase):
    def test_htaccess_formato(self):
        block = build_htaccess_block(_plan_simples(), "2026-06-09")
        self.assertIn(
            "Redirect 301 /pagina-dup/?p=1 https://www.ex.com.br/melhor-pagina/",
            block,
        )

    def test_nginx_formato(self):
        block = build_nginx_block(_plan_simples(), "2026-06-09")
        self.assertIn(
            "location = /pagina-dup/?p=1 { return 301 https://www.ex.com.br/melhor-pagina/; }",
            block,
        )

    def test_blocos_marcados_como_sugestao(self):
        for block in (
            build_htaccess_block(_plan_simples(), "2026-06-09"),
            build_nginx_block(_plan_simples(), "2026-06-09"),
        ):
            self.assertIn("SUGESTAO", block)
            self.assertIn("revisao humana", block.replace("\n# ", " "))

    def test_blocos_sao_ascii(self):
        # Arquivos de config de servidor: nada de acento/emoji
        for block in (
            build_htaccess_block(_plan_simples(), "2026-06-09"),
            build_nginx_block(_plan_simples(), "2026-06-09"),
        ):
            block.encode("ascii")  # não deve levantar

    def test_keyword_com_acento_vira_ascii_no_comentario(self):
        block = build_nginx_block(_plan_simples(), "2026-06-09")
        self.assertIn("# keyword: kw acai", block)


# ---------------------------------------------------------------------------
# Artefatos em disco (storage)
# ---------------------------------------------------------------------------


class TestRedirectArtifacts(unittest.TestCase):
    def setUp(self):
        self._original_dir = storage.RELATORIOS_DIR
        storage.RELATORIOS_DIR = tempfile.mkdtemp(prefix="gsc_test_redir_")

    def tearDown(self):
        shutil.rmtree(storage.RELATORIOS_DIR, ignore_errors=True)
        storage.RELATORIOS_DIR = self._original_dir

    def test_csv_colunas_e_dados(self):
        path = storage.save_redirects_csv("ex.com", "2026-06-09", _plan_simples())

        self.assertTrue(path.endswith("2026-06-09_redirects.csv"))
        with open(path, encoding="utf-8-sig") as f:
            lines = f.read().splitlines()

        # 1ª linha: aviso de sugestão; 2ª: header; 3ª: dados
        self.assertTrue(lines[0].startswith("#"))
        self.assertIn("SUGESTAO", lines[0])

        rows = list(csv.reader(lines[1:]))
        self.assertEqual(
            rows[0], ["from_url", "to_url", "keyword", "severity", "clicks_from", "clicks_to"]
        )
        dado = dict(zip(rows[0], rows[1]))
        self.assertEqual(dado["from_url"], "https://www.ex.com.br/pagina-dup/?p=1")
        self.assertEqual(dado["to_url"], "https://www.ex.com.br/melhor-pagina/")
        self.assertEqual(dado["keyword"], "kw açaí")
        self.assertEqual(dado["severity"], "média")
        self.assertEqual(dado["clicks_from"], "5")
        self.assertEqual(dado["clicks_to"], "50")

    def test_txt_apache_e_nginx_gravados(self):
        plan = _plan_simples()
        pa, pn = storage.save_redirects_txt(
            "ex.com",
            "2026-06-09",
            build_htaccess_block(plan, "2026-06-09"),
            build_nginx_block(plan, "2026-06-09"),
        )
        self.assertTrue(pa.endswith("2026-06-09_redirects_apache.txt"))
        self.assertTrue(pn.endswith("2026-06-09_redirects_nginx.txt"))
        with open(pa, encoding="utf-8") as f:
            self.assertIn("Redirect 301", f.read())
        with open(pn, encoding="utf-8") as f:
            self.assertIn("return 301", f.read())


# ---------------------------------------------------------------------------
# Integração leve com os reporters (sem rede)
# ---------------------------------------------------------------------------


class TestReporterSurfaces(unittest.TestCase):
    def test_dashboard_secao_marcada_como_sugestao(self):
        from reporters.html_reporter import _sec_plano_301

        html_out = _sec_plano_301(_plan_simples())
        self.assertIn("SUGESTÃO", html_out)
        self.assertIn("plano301", html_out)
        self.assertIn("melhor-pagina", html_out)

    def test_dashboard_secao_vazia_sem_redirects(self):
        from reporters.html_reporter import _sec_plano_301

        self.assertEqual(_sec_plano_301({"redirects": []}), "")
        self.assertEqual(_sec_plano_301(None), "")

    def test_dashboard_escapa_keyword_maliciosa(self):
        from reporters.html_reporter import _sec_plano_301

        g = _group(
            query="<script>alert(1)</script>",
            urls=[
                _url({"url": "https://ex.com/a", "clicks": 50}),
                _url({"url": "https://ex.com/b", "clicks": 5}),
            ],
        )
        html_out = _sec_plano_301(build_consolidation_plan([g]))
        self.assertNotIn("<script>alert", html_out)
        self.assertIn("&lt;script&gt;", html_out)

    def test_excel_sheet_plano_301(self):
        from openpyxl import Workbook
        from reporters.excel_reporter import _build_sheet_plano_301

        wb = Workbook()
        ws = wb.active
        data = {"start_date": "2026-05-10", "end_date": "2026-06-08"}
        _build_sheet_plano_301(ws, _plan_simples(), "ex.com", data)

        self.assertEqual(ws.title, "Plano 301")
        self.assertIn("SUGESTÃO", ws["A1"].value)
        # Header na linha 3 e dados na 4
        self.assertEqual(ws.cell(row=3, column=3).value, "Redirect de (301)")
        self.assertEqual(ws.cell(row=4, column=3).value, "https://www.ex.com.br/pagina-dup/?p=1")
        self.assertEqual(ws.cell(row=4, column=4).value, "https://www.ex.com.br/melhor-pagina/")


if __name__ == "__main__":
    unittest.main()
