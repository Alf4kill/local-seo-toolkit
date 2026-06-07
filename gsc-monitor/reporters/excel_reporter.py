"""
excel_reporter.py — Gera relatório Excel (.xlsx) com:
  - Sheet "Resumo"           : visão executiva do site
  - Sheet "URLs por Faixa"   : todas as URLs classificadas por faixa de posição
  - Sheet "Oportunidades CTR": URLs na página 1 com score de CTR perdido

Requer: openpyxl  (pip install openpyxl)
"""

from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Benchmarks de CTR esperado por posição no Google (média de mercado)
# Modifique aqui se quiser usar referências diferentes
# ---------------------------------------------------------------------------
CTR_BENCHMARK = {
    1:  28.5,
    2:  15.7,
    3:  11.0,
    4:   8.0,
    5:   7.2,
    6:   5.1,
    7:   4.0,
    8:   3.2,
    9:   2.8,
    10:  2.5,
}
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Paleta de cores por faixa de posição
# ---------------------------------------------------------------------------
RANGE_COLORS = {
    "Top 3":       "92D050",   # Verde
    "1ª Página":   "E2EFDA",   # Verde claro
    "2ª Página":   "FFEB9C",   # Amarelo
    "3ª Página":   "FFCC99",   # Laranja
    "4ª+ Página":  "FFC7CE",   # Vermelho claro
    "Sem Dados":   "D9D9D9",   # Cinza
}

HEADER_BG    = "2F75B6"   # Azul escuro — cabeçalhos de coluna
HEADER_FG    = "FFFFFF"   # Branco — texto dos cabeçalhos
TITLE_BG     = "1F4E79"   # Azul marinho — título da sheet
SECTION_BG   = "BDD7EE"   # Azul claro — separadores de seção

OPP_HIGH_BG  = "FF7B7B"   # Vermelho claro — score alto  (≥ 50 cliques perdidos)
OPP_MID_BG   = "FFDA6A"   # Amarelo         — score médio (≥ 10)
OPP_LOW_BG   = "E2EFDA"   # Verde claro     — score baixo

HEALTH_GRADE_COLORS = {
    "Excelente": "92D050",   # verde
    "Bom":       "C6EFCE",   # verde claro
    "Regular":   "FFEB9C",   # amarelo
    "Crítico":   "FFC7CE",   # vermelho claro
}
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _set_header(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill      = _fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


def _col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


# ---------------------------------------------------------------------------
# Classificação de posição
# ---------------------------------------------------------------------------

def _classify(position) -> tuple[str, int]:
    """Retorna (label_faixa, ordem_sort)."""
    if position is None:    return "Sem Dados",    6
    if position <= 3:       return "Top 3",        1
    if position <= 10:      return "1ª Página",    2
    if position <= 20:      return "2ª Página",    3
    if position <= 50:      return "3ª Página",    4
    return                         "4ª+ Página",   5


ACOES = {
    "Top 3":       "Proteger posição",
    "1ª Página":   "Otimizar title e meta description",
    "2ª Página":   "Fortalecer conteúdo e links internos",
    "3ª Página":   "Revisar e ampliar conteúdo",
    "4ª+ Página":  "Reestruturar ou consolidar página",
    "Sem Dados":   "Sem impressões no período analisado",
}


# ---------------------------------------------------------------------------
# Score de oportunidade de CTR
# ---------------------------------------------------------------------------

def _expected_ctr(position: float) -> float | None:
    """
    Interpola o CTR esperado (%) para posições decimais.
    Retorna None se posição > 10.
    """
    if position is None or position > 10:
        return None
    pos_floor = max(1, min(int(position), 10))
    pos_ceil  = min(pos_floor + 1, 10)
    frac      = position - int(position)
    ctr_low   = CTR_BENCHMARK.get(pos_floor, CTR_BENCHMARK[10])
    ctr_high  = CTR_BENCHMARK.get(pos_ceil,  CTR_BENCHMARK[10])
    return ctr_low + frac * (ctr_high - ctr_low)


def _opportunity_score(row: dict) -> float:
    """
    Cliques estimados perdidos vs. benchmark de CTR.
    score = impressoes × max(0, ctr_esperado − ctr_real) / 100
    """
    if not row["has_data"] or row["position"] is None or row["position"] > 10:
        return 0.0
    exp = _expected_ctr(row["position"])
    if exp is None:
        return 0.0
    return round(row["impressions"] * max(0.0, exp - row["ctr"]) / 100, 1)


# ---------------------------------------------------------------------------
# Sheet 1 — Resumo executivo
# ---------------------------------------------------------------------------

def _build_sheet_resumo(
    ws,
    report: dict,
    data: dict,
    health: "dict | None" = None,
    kg_result: "dict | None" = None,
) -> None:
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    summary = report["summary"]
    rows    = report["urls"]

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "RELATÓRIO DE POSICIONAMENTO — PRST Monitor"
    t.font      = Font(bold=True, color=HEADER_FG, size=13)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 22)

    # ── Informações gerais ───────────────────────────────────────────────────
    def section(row, label):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value     = label
        c.font      = Font(bold=True, color="1F4E79", size=10)
        c.fill      = _fill(SECTION_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        _row_height(ws, row, 18)

    def info_row(row, label, value, fmt=None):
        a = ws.cell(row=row, column=1, value=label)
        b = ws.cell(row=row, column=2, value=value)
        a.font      = Font(bold=True, size=10)
        a.alignment = Alignment(indent=2)
        b.alignment = Alignment(horizontal="left", indent=1)
        b.font      = Font(size=10)
        if fmt:
            b.number_format = fmt
        _row_height(ws, row, 16)

    section(2, "  Identificação")
    info_row(3,  "Site",             report["site"])
    info_row(4,  "Data do relatório",report["date"])
    info_row(5,  "Período analisado",f"{data['start_date']}  até  {data['end_date']}")
    info_row(6,  "País",             "Global (sem filtro)")

    section(8, "  Visão Geral")
    info_row(9,  "Total de URLs no sitemap",  summary["total_urls_sitemap"])
    info_row(10, "URLs com dados no GSC",      summary["urls_with_data"])
    info_row(11, "URLs sem impressões",         summary["urls_no_impressions"])
    info_row(12, "Posição média do site",       summary["avg_position_site"])
    info_row(13, "Total de cliques (período)",  summary["total_clicks"])
    info_row(14, "Total de impressões",         summary["total_impressions"])
    info_row(15, "CTR médio (%)",               summary["avg_ctr_percent"])

    # ── Distribuição por faixa ───────────────────────────────────────────────
    section(17, "  Distribuição por Faixa de Posição")

    _set_header(ws, 18, 1, "Faixa")
    _set_header(ws, 18, 2, "Quantidade de URLs")
    _set_header(ws, 18, 3, "% do sitemap")
    _row_height(ws, 18, 18)

    total = len(rows)
    range_counts = Counter(_classify(r["position"])[0] for r in rows)
    faixas_order = ["Top 3", "1ª Página", "2ª Página", "3ª Página", "4ª+ Página", "Sem Dados"]

    for i, faixa in enumerate(faixas_order):
        r_idx = 19 + i
        count = range_counts.get(faixa, 0)
        pct   = round(count / total * 100, 1) if total else 0.0
        fill  = _fill(RANGE_COLORS[faixa])
        brd   = _border()

        for col, val in enumerate([faixa, count, pct / 100], start=1):
            cell           = ws.cell(row=r_idx, column=col, value=val)
            cell.fill      = fill
            cell.border    = brd
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=10)
        ws.cell(row=r_idx, column=3).number_format = "0.0%"
        _row_height(ws, r_idx, 16)

    # ── Score de Saúde do Site (opcional — Fase 4d) ─────────────────────────
    if health is not None:
        brd_h   = _border()
        grade   = health["grade"]
        comp    = health["components"]
        g_color = HEALTH_GRADE_COLORS.get(grade, "D9D9D9")

        section(26, "  Saúde do Site")

        # Linha de score — label | valor | grade (fundo colorido)
        a27 = ws.cell(row=27, column=1, value="Score Geral")
        b27 = ws.cell(row=27, column=2, value=health["score"])
        c27 = ws.cell(row=27, column=3, value=grade)
        a27.font = Font(bold=True, size=11)
        a27.alignment = Alignment(indent=2)
        b27.font = Font(bold=True, size=13)
        b27.alignment = Alignment(horizontal="center", vertical="center")
        b27.number_format = "0.0"
        c27.fill = _fill(g_color)
        c27.font = Font(bold=True, size=11)
        c27.alignment = Alignment(horizontal="center", vertical="center")
        for cell in (a27, b27, c27):
            cell.border = brd_h
        _row_height(ws, 27, 24)

        # Componentes
        idx_val = comp["indexation"] if comp["indexation"] is not None else "s/d"
        info_row(28, "Comp. Indexação (%, peso 40%)",    idx_val)
        info_row(29, "Comp. Posicionamento (peso 40%)",   comp["position"])
        info_row(30, "Comp. CTR vs benchmark (peso 20%)", comp["ctr"])

        if not health["has_indexation_data"]:
            note = ws.cell(row=31, column=1,
                           value="* Indexação não executada — score baseado só em Posição + CTR (pesos re-normalizados)")
            note.font      = Font(italic=True, size=8, color="888888")
            note.alignment = Alignment(indent=2)
            ws.merge_cells("A31:C31")
            _row_height(ws, 31, 14)

    # ── Knowledge Graph (opcional — Fase 5a) ────────────────────────────────
    if kg_result is not None:
        if health is not None:
            kg_start = 32 if health.get("has_indexation_data", True) else 33
        else:
            kg_start = 26

        section(kg_start, "  Knowledge Graph")

        if kg_result.get("found"):
            types_str = ", ".join(t for t in kg_result.get("types", []) if t != "Thing") or "–"
            info_row(kg_start + 1, "Entidade encontrada",    "Sim")
            info_row(kg_start + 2, "Nome no KG",             kg_result.get("name", ""))
            info_row(kg_start + 3, "Tipo(s)",                types_str)
            if kg_result.get("description"):
                info_row(kg_start + 4, "Descrição",          kg_result["description"])
            info_row(kg_start + 5, "Score KG",               kg_result.get("score", 0))
        else:
            info_row(kg_start + 1, "Entidade encontrada",    "Não")
            info_row(kg_start + 2, "Observação",
                     "Marca não registrada no Google Knowledge Graph.")

    # ── Larguras ─────────────────────────────────────────────────────────────
    _col_width(ws, 1, 32)
    _col_width(ws, 2, 22)
    _col_width(ws, 3, 16)


# ---------------------------------------------------------------------------
# Sheet 2 — URLs por Faixa de Posição
# ---------------------------------------------------------------------------

def _build_sheet_urls(ws, rows: list[dict], domain: str, data: dict) -> None:
    ws.title = "URLs por Faixa"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"URLs por Faixa de Posição  |  {domain}  |  {data['start_date']} a {data['end_date']}"
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    # Cabeçalhos
    headers = ["Faixa", "Posição", "Cliques", "Impressões", "CTR (%)", "Ação Sugerida", "URL"]
    for col, h in enumerate(headers, start=1):
        _set_header(ws, 2, col, h)
    _row_height(ws, 2, 30)

    # Ordena por faixa → posição
    sorted_rows = sorted(
        rows,
        key=lambda r: (_classify(r["position"])[1], r["position"] or 9999)
    )

    brd = _border()
    for i, r in enumerate(sorted_rows, start=3):
        faixa, _ = _classify(r["position"])
        fill      = _fill(RANGE_COLORS[faixa])

        values = [
            faixa,
            r["position"],
            r["clicks"],
            r["impressions"],
            r["ctr"] / 100 if r["ctr"] else None,   # decimal para formatação %
            ACOES[faixa],
            r["url"],
        ]
        fmts = [None, "0.0", "#,##0", "#,##0", "0.00%", None, None]

        for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
            cell              = ws.cell(row=i, column=col, value=val)
            cell.fill         = fill
            cell.border       = brd
            cell.alignment    = Alignment(vertical="center", wrap_text=(col == 7))
            cell.font         = Font(size=9)
            if fmt:
                cell.number_format = fmt

        _row_height(ws, i, 15)

    # Larguras
    widths = [14, 10, 10, 13, 10, 30, 70]
    for col, w in enumerate(widths, start=1):
        _col_width(ws, col, w)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{len(sorted_rows) + 2}"


# ---------------------------------------------------------------------------
# Sheet 3 — Oportunidades de CTR
# ---------------------------------------------------------------------------

def _build_sheet_oportunidades(
    ws,
    rows: "list[dict]",
    domain: str,
    data: dict,
    nlp_results: "dict | None" = None,
) -> None:
    ws.title = "Oportunidades CTR"
    ws.sheet_view.showGridLines = False

    n_cols     = 11 if nlp_results is not None else 9
    last_col_l = get_column_letter(n_cols)

    # Título
    ws.merge_cells(f"A1:{last_col_l}1")
    t = ws["A1"]
    t.value     = f"Oportunidades de CTR — 1ª Página (posição ≤ 10)  |  {domain}"
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    # Linha de legenda
    ws.merge_cells(f"A2:{last_col_l}2")
    leg = ws["A2"]
    leg.value     = "Score = cliques estimados perdidos vs. benchmark de CTR por posição. Quanto maior, maior o impacto de melhorar o title/meta description."
    leg.font      = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _row_height(ws, 2, 16)

    # Cabeçalhos
    headers = [
        "Posição", "Faixa", "Impressões", "Cliques Reais",
        "CTR Real (%)", "CTR Esperado (%)", "Cliques Perdidos",
        "Score Oportunidade", "URL",
    ]
    if nlp_results is not None:
        headers.append("Entidades Principais")
        headers.append("Categoria NLP")
    for col, h in enumerate(headers, start=1):
        _set_header(ws, 3, col, h)
    _row_height(ws, 3, 30)

    # Monta e filtra linhas de oportunidade
    opp_rows = []
    for r in rows:
        if not r["has_data"] or r["position"] is None or r["position"] > 10:
            continue
        exp_ctr = _expected_ctr(r["position"])
        if exp_ctr is None:
            continue
        score        = _opportunity_score(r)
        cliques_perd = round(r["impressions"] * max(0.0, exp_ctr - r["ctr"]) / 100, 1)
        faixa, _     = _classify(r["position"])
        opp_rows.append({**r, "faixa": faixa, "exp_ctr": exp_ctr, "cliques_perdidos": cliques_perd, "score": score})

    # Ordena por score decrescente
    opp_rows.sort(key=lambda r: r["score"], reverse=True)

    brd = _border()
    for i, r in enumerate(opp_rows, start=4):
        # Cor de fundo da linha por score
        if r["score"] >= 50:
            fill = _fill(OPP_HIGH_BG)
        elif r["score"] >= 10:
            fill = _fill(OPP_MID_BG)
        else:
            fill = _fill(OPP_LOW_BG)

        values = [
            r["position"],
            r["faixa"],
            r["impressions"],
            r["clicks"],
            r["ctr"] / 100 if r["ctr"] else 0.0,
            r["exp_ctr"] / 100,
            r["cliques_perdidos"],
            r["score"],
            r["url"],
        ]
        fmts = ["0.0", None, "#,##0", "#,##0", "0.00%", "0.00%", "#,##0.0", "#,##0.0", None]

        if nlp_results is not None:
            nlp_data   = nlp_results.get(r["url"], {})
            # backward compat: antigo formato era lista direta
            if isinstance(nlp_data, list):
                entities, categories = nlp_data, []
            else:
                entities   = nlp_data.get("entities", [])
                categories = nlp_data.get("categories", [])
            ent_str = ", ".join(e["name"] for e in entities[:4]) if entities else "–"
            if categories:
                cat_label = categories[0]["name"].rsplit("/", 1)[-1]
                cat_str   = f"{cat_label} ({categories[0]['confidence']:.0%})"
            else:
                cat_str = "–"
            values.append(ent_str)
            values.append(cat_str)
            fmts.extend([None, None])

        for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
            cell           = ws.cell(row=i, column=col, value=val)
            cell.border    = brd
            cell.font      = Font(size=9)
            if col == 2:
                cell.fill      = _fill(RANGE_COLORS.get(r["faixa"], "D9D9D9"))
                cell.font      = Font(size=9, bold=True)
                cell.alignment = Alignment(vertical="center", horizontal="center")
            else:
                cell.fill      = fill
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col < 9 else "left",
                    wrap_text=(col >= 9),
                )
            if fmt:
                cell.number_format = fmt

        _row_height(ws, i, 15)

    # Legenda de cores
    legend_row = len(opp_rows) + 6
    ws.cell(row=legend_row, column=1, value="Legenda:").font = Font(bold=True, size=9)
    items = [
        (OPP_HIGH_BG, "Score ≥ 50 — Alto impacto: prioridade máxima"),
        (OPP_MID_BG,  "Score 10–49 — Médio impacto: otimização recomendada"),
        (OPP_LOW_BG,  "Score < 10  — Baixo impacto: monitorar"),
    ]
    for j, (color, label) in enumerate(items):
        r_idx = legend_row + 1 + j
        cell_a = ws.cell(row=r_idx, column=1, value="")
        cell_b = ws.cell(row=r_idx, column=2, value=label)
        cell_a.fill = _fill(color)
        cell_b.font = Font(size=9)
        _row_height(ws, r_idx, 14)

    # Larguras
    widths = [10, 12, 13, 14, 14, 16, 17, 19, 55]
    if nlp_results is not None:
        widths.extend([35, 28])
    for col, w in enumerate(widths, start=1):
        _col_width(ws, col, w)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col_l}{len(opp_rows) + 3}"


# ---------------------------------------------------------------------------
# Sheet — Tendências (Fase 5b)
# ---------------------------------------------------------------------------

TREND_COLORS = {
    "rising":   "92D050",   # verde
    "stable":   "FFEB9C",   # amarelo
    "declining": "FFC7CE",  # vermelho claro
}
TREND_LABELS = {
    "rising":   "↑ Crescente",
    "stable":   "→ Estável",
    "declining": "↓ Em queda",
}


def _build_sheet_trends(
    ws,
    trends_data: dict,
    kw_positions: dict,
    domain: str,
    data: dict,
) -> None:
    """
    trends_data   — {keyword: {"trend", "peak", "latest", "values"}}
    kw_positions  — {keyword: {"position": float, "impressions": int}}
    """
    ws.title = "Trends"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = f"Tendências Google Trends (12 meses)  |  {domain}"
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    ws.merge_cells("A2:F2")
    leg = ws["A2"]
    leg.value     = "Interesse relativo 0–100 no Google Trends. Tendência calculada sobre a média dos primeiros 3 vs. últimos 3 meses."
    leg.font      = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _row_height(ws, 2, 16)

    headers = ["Keyword", "Posição Atual", "Tendência (12m)", "Pico", "Atual", "Período"]
    for col, h in enumerate(headers, start=1):
        _set_header(ws, 3, col, h)
    _row_height(ws, 3, 22)

    brd = _border()
    for i, (kw, td) in enumerate(trends_data.items(), start=4):
        trend       = td.get("trend", "stable")
        fill_color  = TREND_COLORS.get(trend, "D9D9D9")
        trend_label = TREND_LABELS.get(trend, "→ Estável")
        kw_pos      = kw_positions.get(kw, {})
        pos_val     = kw_pos.get("position")

        values_list = [
            kw,
            pos_val,
            trend_label,
            td.get("peak", 0),
            td.get("latest", 0),
            f"{data['start_date']} a {data['end_date']}",
        ]
        fmts = [None, "0.0", None, "0", "0", None]

        for col, (val, fmt) in enumerate(zip(values_list, fmts), start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", horizontal="center" if col > 1 else "left")
            if col == 3:
                cell.fill = _fill(fill_color)
                cell.font = Font(size=9, bold=True)
            else:
                cell.fill = _fill("FFFFFF" if i % 2 == 0 else "F5F5F5")
            if fmt:
                cell.number_format = fmt
        _row_height(ws, i, 16)

    widths = [38, 14, 18, 8, 8, 26]
    for col, w in enumerate(widths, start=1):
        _col_width(ws, col, w)

    ws.freeze_panes = "A4"
    if trends_data:
        ws.auto_filter.ref = f"A3:F{len(trends_data) + 3}"


# ---------------------------------------------------------------------------
# Sheet 4 — Páginas Órfãs  (Fase 4a)
# ---------------------------------------------------------------------------

def _build_sheet_orfas(ws, orphans: list, domain: str, data: dict) -> None:
    ws.title = "Sem Impressões"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value     = (
        f"Páginas sem impressões  |  {domain}  |  "
        f"{data['start_date']} a {data['end_date']}"
    )
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    ws.merge_cells("A2:B2")
    leg = ws["A2"]
    leg.value     = (
        f"Total: {len(orphans)} URL(s) com 0 impressões. "
        "Revisar conteúdo, melhorar links internos ou consolidar."
    )
    leg.font      = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _row_height(ws, 2, 16)

    _set_header(ws, 3, 1, "URL")
    _set_header(ws, 3, 2, "Sugestão")
    _row_height(ws, 3, 22)

    brd       = _border()
    alt_fills = ["FFFFFF", "F5F5F5"]
    for i, entry in enumerate(orphans, start=4):
        fill_c = alt_fills[(i - 4) % 2]
        for col, val in enumerate([entry["url"], entry["suggestion"]], start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = _fill(fill_c)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 1))
        _row_height(ws, i, 15)

    _col_width(ws, 1, 70)
    _col_width(ws, 2, 34)
    ws.freeze_panes = "A4"
    if orphans:
        ws.auto_filter.ref = f"A3:B{len(orphans) + 3}"


# ---------------------------------------------------------------------------
# Sheet 5 — Histórico de Posição  (Fase 4c)
# ---------------------------------------------------------------------------

def _build_sheet_historico(ws, historico: dict, domain: str) -> None:
    ws.title = "Histórico"
    ws.sheet_view.showGridLines = False

    snapshots = historico.get("snapshots", [])
    recent    = snapshots[-8:]   # últimos 8 snapshots

    if not recent:
        ws.cell(row=1, column=1, value="Sem dados históricos disponíveis.")
        return

    dates    = [s["date"] for s in recent]
    n_dates  = len(dates)
    n_cols   = n_dates + 2   # URL + n datas + Tendência
    last_col = get_column_letter(n_cols)

    # Coleta todas as URLs presentes em qualquer snapshot
    all_urls = sorted({url for s in recent for url in s["urls"]})

    # Título
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = f"Histórico de Posicionamento  |  {domain}  |  últimos {n_dates} snapshots"
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    # Cabeçalhos
    _set_header(ws, 2, 1, "URL")
    for i, date_str in enumerate(dates):
        _set_header(ws, 2, i + 2, f"Pos.\n{date_str}")
    _set_header(ws, 2, n_dates + 2, "Tendência")
    _row_height(ws, 2, 36)

    brd = _border()
    for r_idx, url in enumerate(all_urls, start=3):
        positions = [s["urls"].get(url, {}).get("position") for s in recent]

        # Cálculo de tendência (primeira vs. última posição disponível)
        first_pos = next((p for p in positions if p is not None), None)
        last_pos  = next((p for p in reversed(positions) if p is not None), None)

        if first_pos is not None and last_pos is not None and first_pos != last_pos:
            delta = first_pos - last_pos   # positivo = melhoria (posição menor)
            if abs(delta) >= 2.0:
                trend       = f"{'↑' if delta > 0 else '↓'} {abs(delta):.1f}"
                trend_color = "92D050" if delta > 0 else "FFC7CE"
            else:
                trend, trend_color = "→", "D9D9D9"
        elif first_pos is None and last_pos is None:
            trend, trend_color = "s/d", "D9D9D9"
        else:
            trend, trend_color = "→", "D9D9D9"

        # Coluna URL
        url_c           = ws.cell(row=r_idx, column=1, value=url)
        url_c.font      = Font(size=8)
        url_c.border    = brd
        url_c.alignment = Alignment(vertical="center")

        # Colunas de posição por data
        for i, pos in enumerate(positions):
            cell = ws.cell(row=r_idx, column=i + 2, value=pos)
            faixa = _classify(pos)[0] if pos is not None else "Sem Dados"
            cell.fill         = _fill(RANGE_COLORS[faixa])
            cell.border       = brd
            cell.font         = Font(size=9)
            cell.alignment    = Alignment(horizontal="center", vertical="center")
            if pos is not None:
                cell.number_format = "0.0"

        # Coluna tendência
        trend_c = ws.cell(row=r_idx, column=n_dates + 2, value=trend)
        trend_c.fill      = _fill(trend_color)
        trend_c.border    = brd
        trend_c.font      = Font(size=10, bold=True)
        trend_c.alignment = Alignment(horizontal="center", vertical="center")

        _row_height(ws, r_idx, 15)

    # Larguras
    _col_width(ws, 1, 62)
    for i in range(n_dates):
        _col_width(ws, i + 2, 14)
    _col_width(ws, n_dates + 2, 12)

    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Sheet 6 — Canibalização de Keywords  (Fase 4b)
# ---------------------------------------------------------------------------

def _build_sheet_canibalizacao(ws, cannibalization: list, domain: str, data: dict) -> None:
    ws.title = "Canibalização"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = (
        f"Canibalização de Keywords  |  {domain}  |  "
        f"{data['start_date']} a {data['end_date']}"
    )
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    ws.merge_cells("A2:G2")
    leg = ws["A2"]
    leg.value     = (
        "Keywords onde 2+ páginas do mesmo site competem no Google (volume e posição relevantes). "
        "Cor da keyword = severidade (impressões em disputa). Consolidar ou diferenciar o conteúdo."
    )
    leg.font      = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _row_height(ws, 2, 16)

    headers = ["Keyword", "URL", "Posição", "Cliques", "Impressões", "CTR (%)", "Severidade / Sugestão"]
    for col, h in enumerate(headers, start=1):
        _set_header(ws, 3, col, h)
    _row_height(ws, 3, 22)

    # Cor da célula da keyword = severidade (impressões em disputa entre as URLs)
    SEVERITY_FILL  = {"alta": "FFC7CE", "média": "FFEB9C", "baixa": "C6EFCE"}
    SEVERITY_LABEL = {"alta": "ALTA", "média": "MÉDIA", "baixa": "BAIXA"}

    group_fills = ["EBF3FB", "F2F7EE"]   # azul claro / verde claro alternando
    brd     = _border()
    row_idx = 4

    for g_idx, group in enumerate(cannibalization):
        bg_color = group_fills[g_idx % 2]
        sev      = group.get("severity", "baixa")
        for u_idx, u in enumerate(group["urls"]):
            keyword_val = group["query"] if u_idx == 0 else ""
            sugestao    = (
                f"Severidade {SEVERITY_LABEL.get(sev, sev.upper())} — consolidar ou diferenciar"
                if u_idx == 0 else ""
            )

            values = [
                keyword_val,
                u["url"],
                u["position"],
                u["clicks"],
                u["impressions"],
                u["ctr"] / 100 if u.get("ctr") else 0.0,
                sugestao,
            ]
            fmts = [None, None, "0.0", "#,##0", "#,##0", "0.00%", None]

            for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                # Keyword (1ª linha do grupo) recebe a cor da severidade; demais células, a do grupo
                if col == 1 and u_idx == 0:
                    cell.fill = _fill(SEVERITY_FILL.get(sev, "D9D9D9"))
                else:
                    cell.fill = _fill(bg_color)
                cell.border    = brd
                cell.font      = Font(size=9, bold=(col == 1 and u_idx == 0))
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col in (3, 4, 5, 6) else "left",
                    wrap_text=(col == 2),
                )
                if fmt:
                    cell.number_format = fmt
            _row_height(ws, row_idx, 15)
            row_idx += 1

        # Linha separadora entre grupos
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).fill = _fill("EEEEEE")
        _row_height(ws, row_idx, 4)
        row_idx += 1

    widths = [28, 55, 10, 10, 12, 10, 28]
    for col, w in enumerate(widths, start=1):
        _col_width(ws, col, w)

    ws.freeze_panes = "A4"
    if cannibalization:
        ws.auto_filter.ref = f"A3:G{row_idx - 1}"


# ---------------------------------------------------------------------------
# Sheet — Qualidade de Conteúdo  (Move 1)
# ---------------------------------------------------------------------------

CQ_VERDICT_FILL = {
    "ok":             "C6EFCE",
    "atencao":        "FFEB9C",
    "over_otimizado": "FFC7CE",
    "raso":           "FFC7CE",
}


def _build_sheet_content_quality(ws, content_results: dict, domain: str) -> None:
    ws.title = "Qualidade Conteúdo"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = f"Qualidade de Conteúdo (heurística)  |  {domain}"
    t.font      = Font(bold=True, color=HEADER_FG, size=11)
    t.fill      = _fill(TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    _row_height(ws, 1, 20)

    ws.merge_cells("A2:H2")
    leg = ws["A2"]
    leg.value     = (
        "Diagnóstico de over-optimization / conteúdo raso nas páginas de oportunidade. "
        "Heurística correlacionada com o que os core updates valorizam — não é o algoritmo do Google."
    )
    leg.font      = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _row_height(ws, 2, 16)

    headers = ["Veredito", "URL", "Palavras", "Densidade (%)", "Keyword-alvo",
               "Diversidade", "Entidades", "Alertas"]
    for col, h in enumerate(headers, start=1):
        _set_header(ws, 3, col, h)
    _row_height(ws, 3, 24)

    brd   = _border()
    order = {"raso": 0, "over_otimizado": 1, "atencao": 2, "ok": 3}
    items = sorted(content_results.items(),
                   key=lambda kv: order.get(kv[1].get("verdict", "ok"), 9))

    for i, (url, cq) in enumerate(items, start=4):
        verdict = cq.get("verdict", "ok")
        fill_v  = _fill(CQ_VERDICT_FILL.get(verdict, "D9D9D9"))
        ent     = cq.get("entity_count")
        values = [
            cq.get("verdict_label", verdict),
            url,
            cq.get("word_count", 0),
            (cq.get("keyword_density", 0) or 0) / 100,   # decimal p/ formatação %
            cq.get("densest_keyword") or "—",
            cq.get("vocab_diversity", 0),
            ent if ent is not None else "s/d",
            " · ".join(cq.get("reasons", [])) or "—",
        ]
        fmts = [None, None, "#,##0", "0.0%", None, "0.00", None, None]
        for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border    = brd
            cell.font      = Font(size=9, bold=(col == 1))
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col in (3, 4, 6, 7) else "left",
                wrap_text=(col in (2, 8)),
            )
            cell.fill = fill_v if col == 1 else _fill("FFFFFF" if i % 2 == 0 else "F5F5F5")
            if fmt:
                cell.number_format = fmt
        _row_height(ws, i, 28)

    widths = [22, 50, 10, 13, 26, 12, 10, 60]
    for col, w in enumerate(widths, start=1):
        _col_width(ws, col, w)
    ws.freeze_panes = "A4"
    if items:
        ws.auto_filter.ref = f"A3:H{len(items) + 3}"


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def generate_excel(
    domain: str,
    today: str,
    data: dict,
    report: dict,
    health: "dict | None" = None,
    orphans: "list | None" = None,
    historico_posicao: "dict | None" = None,
    cannibalization: "list | None" = None,
    kg_result: "dict | None" = None,
    trends_data: "dict | None" = None,
    query_rows: "list | None" = None,
    nlp_results: "dict | None" = None,
    content_results: "dict | None" = None,
) -> Workbook:
    """
    Gera o Workbook Excel com até 9 sheets.

    Parâmetros obrigatórios:
        domain  — domínio do site (ex: www.exemplo.com.br)
        today   — data da execução (YYYY-MM-DD)
        data    — resultado bruto de position_fetcher.fetch_positions()
        report  — resultado de position_reporter.build_position_report()

    Parâmetros opcionais (Fase 4):
        health            — dict de analytics.calculate_health_score()  → seção em Resumo
        orphans           — list de analytics.detect_orphan_pages()     → sheet "Páginas Órfãs"
        historico_posicao — dict de storage.load_historico_posicao()    → sheet "Histórico"
        cannibalization   — list de analytics.detect_cannibalization()  → sheet "Canibalização"

    Parâmetros opcionais (Fase 5):
        kg_result         — dict de knowledge_graph.search_entity()     → seção em Resumo
        trends_data       — dict de trends_fetcher.fetch_trends()       → sheet "Trends"
        query_rows        — list de position_fetcher.fetch_query_positions() → posições para Trends
        nlp_results       — dict de nlp_analyzer.analyze_opportunity_urls() → coluna em Oportunidades

    Retorna o objeto Workbook (salvo por storage.save_excel_report).
    """
    wb = Workbook()

    ws_resumo = wb.active
    _build_sheet_resumo(ws_resumo, report, data, health, kg_result)

    ws_urls = wb.create_sheet()
    _build_sheet_urls(ws_urls, report["urls"], domain, data)

    ws_opp = wb.create_sheet()
    _build_sheet_oportunidades(ws_opp, report["urls"], domain, data, nlp_results)

    if trends_data and query_rows:
        # Monta lookup {keyword: {position, impressions}} para a sheet de Trends
        kw_pos: dict = {}
        for r in query_rows:
            q = r.get("query", "")
            if q in trends_data:
                pos = r.get("position", 9999)
                if q not in kw_pos or pos < kw_pos[q]["position"]:
                    kw_pos[q] = {"position": pos, "impressions": r.get("impressions", 0)}
        ws_trends = wb.create_sheet()
        _build_sheet_trends(ws_trends, trends_data, kw_pos, domain, data)

    if orphans:
        ws_orf = wb.create_sheet()
        _build_sheet_orfas(ws_orf, orphans, domain, data)

    if historico_posicao and len(historico_posicao.get("snapshots", [])) >= 2:
        ws_hist = wb.create_sheet()
        _build_sheet_historico(ws_hist, historico_posicao, domain)

    if cannibalization:
        ws_can = wb.create_sheet()
        _build_sheet_canibalizacao(ws_can, cannibalization, domain, data)

    if content_results:
        ws_cq = wb.create_sheet()
        _build_sheet_content_quality(ws_cq, content_results, domain)

    return wb
