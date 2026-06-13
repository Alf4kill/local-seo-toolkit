"""
poda.py — GSC Monitor: Plano de Poda de páginas antigas via terminal.

Premissa (convenção da empresa): o sitemap lista TODAS as páginas ativas.
URLs que o Google conhece e que estão fora do sitemap são páginas antigas —
candidatas a remoção (410/404) ou a 301 quando ainda rendem impressões.

Fluxo em duas etapas (revisão humana obrigatória entre elas):

    # Etapa 1 — gera o plano ({data}_poda.csv, editável)
    python poda.py --site www.exemplo.com.br

    # (analista revisa o CSV: acao_final = 410/404/301/manter, destino_final)

    # Etapa 2 — compila o CSV revisado em blocos de servidor
    python poda.py --site www.exemplo.com.br --compilar
    python poda.py --site www.exemplo.com.br --compilar relatorios/.../2026-06-12_poda.csv

Opções:
    --min-impressoes N   Piso de impressões para marcar "revisar" (padrão: 10)
    --dias N             Janela da Search Analytics (padrão: 480 ≈ 16 meses)
    --importar-gsc ARQ   Export do relatório "Páginas" do GSC (csv/txt/xlsx ou
                         pasta) — cobre URLs sem impressões, invisíveis à API
    --sem-fallback-home  Não sugere a home quando não há destino por query/slug
                         (deixa destino_final em branco em vez do fallback)
    --no-cache           Ignora o cache e força dados frescos da API
"""

import sys

# Força UTF-8 no stdout/stderr (terminais Windows usam cp1252 por padrão).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import argparse
import os
from datetime import date

from core.pruning import (
    PODA_MIN_IMPRESSIONS,
    build_poda_htaccess,
    build_poda_nginx,
    build_poda_php,
    build_poda_redirect,
    build_pruning_plan,
    extract_urls_from_lines,
    parse_plan_csv,
    print_compile_result,
    print_pruning_plan,
)
from core.urls import normalize_domain

# Janela retroativa do Plano de Poda: ~16 meses, o máximo que a Search
# Analytics API guarda. Janela longa de propósito — uma página antiga que
# rendeu impressões em qualquer momento do período aparece no diff.
PODA_DAYS_BACK = 480


def _load_export_urls(paths: list, domain: str) -> list:
    """
    Lê arquivos exportados do relatório de indexação do GSC (download manual
    na UI — a API não expõe esse relatório) e extrai as URLs do domínio.

    Aceita .csv/.txt (texto) e .xlsx (todas as células de texto); um caminho
    de pasta expande para os arquivos suportados dentro dela.
    """
    files = []
    for p in paths:
        if os.path.isdir(p):
            files += [
                os.path.join(p, f)
                for f in sorted(os.listdir(p))
                if f.lower().endswith((".csv", ".txt", ".xlsx"))
            ]
        else:
            files.append(p)

    urls = []
    for fp in files:
        if not os.path.exists(fp):
            print(f"[poda] AVISO: arquivo de export não encontrado: {fp}")
            continue
        if fp.lower().endswith(".xlsx"):
            from openpyxl import load_workbook

            wb = load_workbook(fp, read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    lines.append(" ".join(str(c) for c in row if isinstance(c, str)))
            wb.close()
        else:
            with open(fp, encoding="utf-8-sig", errors="replace") as f:
                lines = f.readlines()
        found = extract_urls_from_lines(lines, domain)
        print(f"[poda] Export importado: {os.path.basename(fp)} — {len(found)} URL(s) do domínio")
        urls += found
    return urls


def run_poda_plan(
    site: str,
    min_impressions: int = PODA_MIN_IMPRESSIONS,
    use_cache: bool = True,
    service=None,
    days_back: int = PODA_DAYS_BACK,
    gsc_exports: "list | None" = None,
    home_fallback: bool = True,
) -> dict:
    """
    Etapa 1: autentica, busca sitemap + Search Analytics (janela longa) e
    gera o Plano de Poda, salvando o CSV editável. Reutilizável pela GUI.

    A Search Analytics só enxerga URLs que APARECERAM na busca no período —
    URLs antigas sem nenhuma impressão são invisíveis à API. Para essas,
    `gsc_exports` aceita arquivos exportados manualmente do relatório
    "Páginas" do GSC (csv/txt/xlsx ou pasta), que entram no plano com
    métricas 0 e origem "export-gsc".

    Retorna {"plan": dict, "csv_path": str | None}.
    Exceções (auth/sitemap/API) propagam para o chamador tratar.
    """
    from core.auth import build_service
    from core.sitemap import fetch_urls
    from core.storage import save_poda_csv
    from fetchers.position_fetcher import fetch_positions, fetch_query_positions

    domain = normalize_domain(site)
    today = date.today().isoformat()

    if service is None:
        print("[poda] Autenticando no Google Search Console...")
        service = build_service()

    print("[poda] Buscando URLs no sitemap (whitelist de páginas ativas)...")
    sitemap_urls = fetch_urls(domain)
    if not sitemap_urls:
        raise RuntimeError("Sitemap vazio — sem whitelist não há como detectar URLs antigas.")

    data = fetch_positions(service, site, sitemap_urls, use_cache=use_cache, days_back=days_back)
    ghost_rows = data.get("ghost_rows", [])

    extra_urls = _load_export_urls(gsc_exports, domain) if gsc_exports else None

    query_rows = None
    if ghost_rows:
        try:
            query_rows = fetch_query_positions(
                service, site, use_cache=use_cache, days_back=days_back
            )
        except Exception as exc:  # noqa: BLE001 — queries são enriquecimento
            print(f"[poda] AVISO: falha ao buscar queries ({exc}) — plano sem contexto de queries.")

    api_pages = {g["url"]: g for g in ghost_rows}
    plan = build_pruning_plan(
        api_pages,
        sitemap_urls,
        query_rows=query_rows,
        min_impressions=min_impressions,
        extra_urls=extra_urls,
        home_fallback=home_fallback,
    )
    print_pruning_plan(plan)

    csv_path = None
    if plan["entries"]:
        csv_path = save_poda_csv(domain, today, plan)
        print("[poda] Revise o CSV (acao_final/destino_final) e rode a etapa 2 (--compilar).")

    return {"plan": plan, "csv_path": csv_path}


def run_poda_compile(site: str, csv_path: "str | None" = None) -> dict:
    """
    Etapa 2: lê o CSV do plano (revisado pelo analista), valida as decisões e
    gera os blocos Apache (RedirectMatch ancorado + estilo Redirect simples)
    /nginx + a versão PHP (para hospedagens onde não dá para editar a config
    do servidor). Reutilizável pela GUI (app_poda.py).

    Retorna {"parsed": dict, "path_apache": str|None, "path_nginx": str|None,
             "path_redirect": str|None, "path_php": str|None, "csv_path": str}.
    """
    from core.storage import (
        latest_poda_csv,
        load_poda_csv_lines,
        save_poda_php,
        save_poda_redirect,
        save_poda_txt,
    )

    domain = normalize_domain(site)
    today = date.today().isoformat()

    if not csv_path:
        csv_path = latest_poda_csv(domain)
        if not csv_path:
            raise RuntimeError(
                f"Nenhum *_poda.csv encontrado para {domain}. Rode a etapa 1 primeiro."
            )
    if not os.path.exists(csv_path):
        raise RuntimeError(f"Arquivo não encontrado: {csv_path}")

    print(f"[poda] Compilando plano revisado: {csv_path}")
    parsed = parse_plan_csv(load_poda_csv_lines(csv_path))
    print_compile_result(parsed)

    path_apache = path_nginx = path_redirect = path_php = None
    if parsed["entries"]:
        path_apache, path_nginx = save_poda_txt(
            domain,
            today,
            build_poda_htaccess(parsed["entries"], today),
            build_poda_nginx(parsed["entries"], today),
        )
        path_redirect = save_poda_redirect(
            domain, today, build_poda_redirect(parsed["entries"], today)
        )
        php_name = f"{today}_poda.php"
        path_php = save_poda_php(
            domain, today, build_poda_php(parsed["entries"], today, filename=php_name)
        )
    else:
        print("[poda] Nenhuma diretiva a compilar (tudo pendente/manter ou com erro).")

    return {
        "parsed": parsed,
        "path_apache": path_apache,
        "path_nginx": path_nginx,
        "path_redirect": path_redirect,
        "path_php": path_php,
        "csv_path": csv_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Plano de Poda — remoção planejada de URLs antigas (fora do sitemap)"
    )
    parser.add_argument(
        "--site",
        required=True,
        help="Domínio a analisar. Ex: www.exemplo.com.br ou sc-domain:exemplo.com.br",
    )
    parser.add_argument(
        "--compilar",
        nargs="?",
        const="",
        default=None,
        metavar="CSV",
        help=(
            "Etapa 2: compila o CSV revisado em blocos Apache/nginx. "
            "Sem valor, usa o *_poda.csv mais recente do domínio."
        ),
    )
    parser.add_argument(
        "--min-impressoes",
        type=int,
        default=PODA_MIN_IMPRESSIONS,
        help=f"Piso de impressões para marcar 'revisar' (padrão: {PODA_MIN_IMPRESSIONS}).",
    )
    parser.add_argument(
        "--dias",
        type=int,
        default=PODA_DAYS_BACK,
        help=(
            f"Janela retroativa em dias da Search Analytics (padrão: {PODA_DAYS_BACK} "
            f"≈ 16 meses, o máximo do GSC)."
        ),
    )
    parser.add_argument(
        "--importar-gsc",
        nargs="+",
        default=None,
        metavar="ARQUIVO",
        help=(
            "Arquivo(s) ou pasta com o export do relatório 'Páginas' do GSC "
            "(csv/txt/xlsx). Cobre URLs antigas sem nenhuma impressão, que a "
            "API não enxerga."
        ),
    )
    parser.add_argument(
        "--no-cache",
        dest="no_cache",
        action="store_true",
        help="Ignora o cache e força novas chamadas à API.",
    )
    parser.add_argument(
        "--sem-fallback-home",
        dest="sem_fallback_home",
        action="store_true",
        help=(
            "Desliga o fallback da home: URLs 'revisar' sem destino por "
            "query/slug ficam com destino_final em branco (em vez de sugerir "
            "a home como ultimo recurso)."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    domain = normalize_domain(args.site)
    print(f"\n=== Plano de Poda — {domain} — {date.today().isoformat()} ===\n")

    try:
        if args.compilar is not None:
            run_poda_compile(args.site, csv_path=args.compilar or None)
        else:
            run_poda_plan(
                args.site,
                min_impressions=args.min_impressoes,
                use_cache=not args.no_cache,
                days_back=args.dias,
                gsc_exports=args.importar_gsc,
                home_fallback=not args.sem_fallback_home,
            )
    except (FileNotFoundError, RuntimeError) as exc:
        print(f"\n[ERRO] {exc}")
        sys.exit(1)

    print("\n[poda] Concluído.")


if __name__ == "__main__":
    main()
